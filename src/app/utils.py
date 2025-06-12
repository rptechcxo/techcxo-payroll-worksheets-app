from typing import List
import holidays
from datetime import date, datetime, timedelta
from openpyxl import Workbook, load_workbook



MONTH = [
    "January",
    "February",
    "March",
    "April",
    "May",
    "June",
    "July",
    "August",
    "September",
    "October",
    "November",
    "December"
]


def worksheet_name(wb: Workbook, expected_sheet_name: str):
    sheet_name_list = [sheet.strip() for sheet in wb.sheetnames]
    for i in range(len(sheet_name_list)):
        if expected_sheet_name == sheet_name_list[i].strip():
            return sheet_name_list[i]

    return None


def is_match(target: str, compare: str, exact_match: bool=False) -> bool:
    if exact_match:
        return target.strip() == compare.strip()
    else:
        return target.strip() in compare.strip()


def find_cell(ws, target_value: str, regex: bool=False, case_sensitive=True, exact_match=False) -> tuple[int, int] | None:
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            if cell.value is None:
                continue
            elif not case_sensitive:
                if is_match(target_value.lower(), str(cell.value).lower(), exact_match):
                    return (cell.row, cell.column)
            elif is_match(target_value, str(cell.value), exact_match):
                return (cell.row, cell.column)

    return None


def get_adjusted_tenth(date_str, country='US', state=None):
    """
    Given a date string, returns the adjusted 10th day of the following month,
    formatted as 'Mon 10th', where the suffix is adjusted properly and the
    month is abbreviated.

    The 10th is chosen from the next month of the given date. If the 10th falls
    on a weekend or a bank holiday (based on country and optional state),
    the function finds the closest previous weekday that is not a holiday.

    Parameters:
    - date_str (str): The input date in 'YYYY-MM-DD' format.
    - country (str): Country code for holidays (default 'US').
    - state (str or None): Optional state code (e.g., 'CA') for localized holidays.

    Returns:
    - str: Adjusted date formatted like 'Feb 10th'.
    """

    def day_suffix(day):
        if 11 <= day <= 13:
            return "th"
        return {1: "st", 2: "nd", 3: "rd"}.get(day % 10, "th")

    # Convert input date
    base_date = datetime.strptime(date_str, "%Y-%m-%d")

    # Move to next month
    if base_date.month == 12:
        target_month = 1
        target_year = base_date.year + 1
    else:
        target_month = base_date.month + 1
        target_year = base_date.year

    # Holiday calendar
    holiday_calendar = holidays.country_holidays(country, years=target_year, subdiv=state)

    # Start at the 10th
    target_date = datetime(target_year, target_month, 10)

    # Adjust backwards for weekends or holidays
    while target_date.weekday() >= 5 or target_date in holiday_calendar:
        target_date -= timedelta(days=1)

    # Format output like "Feb 10th"
    suffix = day_suffix(target_date.day)
    formatted = target_date.strftime(f"%b {target_date.day}{suffix}")
    return formatted


def duplicate_worksheet(wb: Workbook, sheet_name: str, save_path: str):
    """
    Duplicates a worksheet in an Excel workbook.

    Args:
        wb (Workbook): The workbook object.
        sheet_name (str): The name of the sheet to duplicate.

    Returns:
        str: The name of the new duplicated sheet.
    """
    orig_ws = wb[wb.sheetnames[0]]
    ws = wb.copy_worksheet(orig_ws)
    ws.title = sheet_name
    s = find_cell(ws, "December")
    for row in ws.iter_rows(min_row=s[0]+1, max_row=s[0]+23, min_col=s[1], max_col=s[1]+12):
        for cell in row:
            cell.value = None
            cell.comment = None
            cell.number_format = '_(* #,##0.00_);_(* (#,##0.00);_(* "-"??_);_(@_)'
    
    for row in ws.iter_rows(min_row=s[0]-1, max_row=s[0]-1, min_col=s[1], max_col=s[1]+12):
        for cell in row:
            cell.value = None

    wb.move_sheet(wb[sheet_name], offset=-len(wb.sheetnames)+1)
    wb.save(save_path)


def copy_comments(src_path: str, wb: Workbook, partner_name: str, month: str):
    """
    Copies comments from the first worksheet to the new worksheet.

    Args:
        wb (Workbook): The workbook object.
        partner_name (str): The name of the partner.
        columns (List[str]): List of columns to copy comments from.
        month (str): The month for which the comments are being copied.

    Returns:
        None
    """
    
    columns = [
        "Partner Receivable",
        "A/R Backouts",
        "A/R Addbacks",
        "Other Adjustments"
    ] 

    src_wb = load_workbook(src_path)
    src_ws = src_wb['Details']
    dst_ws = wb.active

    src_row = find_cell(src_ws, partner_name, case_sensitive=False)[0]
    dst_col = find_cell(dst_ws, month, exact_match=True)[1]

    for col in columns:
        src_col = find_cell(src_ws, col.split(" ")[-1])[1]
        dst_row = find_cell(dst_ws, col)[0]

        src_comment = src_ws.cell(row=src_row, column=src_col).comment
        dst_ws.cell(row=dst_row, column=dst_col).comment = src_comment
