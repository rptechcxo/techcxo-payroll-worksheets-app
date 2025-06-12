from logging import basicConfig, getLogger, INFO
from app.utils import (
    worksheet_name,
    find_cell,
    MONTH,
    get_adjusted_tenth,
    duplicate_worksheet,
    copy_comments
)
from app.extraction import import_tables
from app.transform import data_to_worksheet
from glob import glob
from os import path
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from pandas import Series
from datetime import datetime, timedelta, date
from gooey import Gooey, GooeyParser
from pandas import DataFrame
import re


logger = getLogger(__name__)
basicConfig(
    level=INFO,
    format='%(asctime)s | %(levelname)s | %(message)s',
    datefmt='%m/%d/%Y %I:%M:%S %p',
)


def to_eom(dt: datetime) -> datetime:
    # Get the first day of the next month
    if dt.month == 12:
        first_of_next = dt.replace(year=dt.year + 1, month=1, day=1)
    else:
        first_of_next = dt.replace(month=dt.month + 1, day=1)

    # Subtract 1 microsecond to get the last moment of the current month
    eom = first_of_next - timedelta(days=1)
    return str(eom)


@Gooey(
    program_name='Payroll Worksheets',
    default_size=(600, 500),
    required_cols=1,
    optional_cols=2,
    body_bg_color="#989898",
    header_bg_color="#989898",
    footer_bg_color="#989898"
)
def main():
    parser = GooeyParser(description="Payroll Worksheets")
    parser.add_argument('output_folder', help="Folder with all the partner worksheets", widget='DirChooser') 
    parser.add_argument('input_file', help="File with the payroll data", widget='FileChooser')
    parser.add_argument('year_month', help="Year and Month to process", widget='DateChooser', default=to_eom(datetime.now()).split(" ")[0]),
    args = parser.parse_args()

    inputs = {
        "data_path": args.input_file,
        "partner_worksheet_folder": args.output_folder,
        "year_month": to_eom(datetime.strptime(args.year_month, "%Y-%m-%d")).split(" ")[0],
    }

    year_month = datetime.strptime(inputs["year_month"], "%Y-%m-%d") + timedelta(days=28)
    month = int(inputs["year_month"].split("-")[1]) - 1
    year = year_month.strftime("%Y")
    details, deductions = import_tables(inputs["data_path"])

    partner_list = glob(path.join(
        inputs["partner_worksheet_folder"],
        "*.xlsx"
    ))
    
    for partner in partner_list:
        logger.info(f"Processing {partner.split('/')[-1]}...")
        try:
            wb = load_workbook(partner)
            sheet_name = worksheet_name(wb, year)
            if sheet_name is None:
                logger.warning(f"Sheet {year} not found in <{path.basename(partner)}>")
                logger.info(f"Creating new Sheet {year} in <{path.basename(partner)}>")
                sheet_name = year
                duplicate_worksheet(wb, sheet_name, partner)

            wb = load_workbook(partner)
            wb.active = wb[sheet_name]
            ws = wb.active

            coord = find_cell(ws, "TechCXO")
            partner_name = ws.cell(row=coord[0]+3, column=coord[1]).value
            partner_name = partner_name.strip().lower()
            partner_name = re.sub(r'\s+', ' ', partner_name)
            employee_code = ws.cell(row=coord[0]+4, column=coord[1]).value

            coord = find_cell(ws, MONTH[month], exact_match=True)
            ws.cell(row=coord[0]-1, column=coord[1]).value = get_adjusted_tenth(inputs["year_month"])
            ws.cell(row=coord[0]-1, column=coord[1]).alignment = Alignment(horizontal="center")

            detail_data = details.loc[partner_name]
            if employee_code in deductions.index:
                deduction_data = deductions.loc[employee_code]
                if isinstance(deduction_data, Series):
                    deduction_data = deduction_data.to_frame().T
                deduction_data = deduction_data.set_index("cdeductcode")
            else:
                deduction_data = DataFrame()

            output_data  = data_to_worksheet(detail_data, deduction_data)
            for i, row in enumerate(range(coord[0]+1, coord[0]+len(output_data)+1)):
                cell = ws.cell(row=row, column=coord[1])
                cell.number_format = '_(* #,##0.00_);_(* (#,##0.00);_(* "-"??_);_(@_)'
                if cell.value is None or str(cell.value).strip() == "":
                    cell.value = output_data[i]
                else:
                    cell.value = output_data[i]
                    logger.warning("Cell %s is not empty, overwriting value", cell.coordinate)
            
            copy_comments(inputs["data_path"], wb, partner_name, MONTH[month])

            wb.save(partner)
            logger.info(f"Finished processing")
        except Exception as e:
            logger.warning(f"Partner file failed to update!")
            logger.warning(f"Failure reason: {e}")

if __name__ == "__main__":
    main()
